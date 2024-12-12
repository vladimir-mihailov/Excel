from openpyxl import load_workbook
workbook = load_workbook(filename="reviews-sample.xlsx")
workbook.sheetnames
['Sheet 1']

sheet = workbook.active
sheet
<Worksheet"Sheet 1">

sheet.title
'Shet 1'
sheet["A1"]
<Cell 'Seet 1', A1>
